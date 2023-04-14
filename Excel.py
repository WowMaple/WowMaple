
import openpyxl

wb_name = '测试.xlsx'

# wb = openpyxl.Workbook()                # 新建工作薄

wb = openpyxl.load_workbook(wb_name)   # 读取工作薄
# wb.save(workbook_name)                 # 保存工作薄
#
# ws1 = wb.active               # 获取活动工作表
# ws2 = wb.worksheets[0]        # 获取第一张工作表，通过索引值，索引值从0开始
# ws3 = wb['Sheet1']            # 通过表名，获取工作表
#
# for wss in wb.worksheets:     # 获取工作薄中的所有工作表
#     print(wss)
#
# ws_name = wb.sheetnames       # 获取工作薄中的所有工作表名，放入一个列表。
# print(ws_name)
#
# wb.worksheets[1].title = '修改了'  # 修改第二张工作表名
# wb.save(wb_name)
#
# for sn in wb.worksheets:                            # 批量修改表名
#     sn.title = sn.title + '批量修改后的表名'
# wb.save('批量修改表名后.xlsx')

# wb.create_sheet('测试一下新建表')
#
# wb.copy_worksheet(wb.worksheets[0]).title = '复制过来的'
#
# wb.remove(wb['表名'])

# wb.save('测试.xlsx')
ws = wb.worksheets[0]
# print(ws.cell(2, 1).value)
# print(ws['a2'].value)
# print(list(ws.values))          # 获取到的是一个嵌套的元组，一个元素是由一行数据组成的元组。
for i in ws['a:b']:       # 循环获取每一行的数据
    for j in i:             # 循环获取一行中的每一个单元格
        j.value = '不是吧'

wb.save(wb_name)

